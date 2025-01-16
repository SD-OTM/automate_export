import pandas as pd
from openpyxl import Workbook

# Load the original Excel file
file_path = "Book1.xlsx"  # Replace with the path to your Excel file
df = pd.read_excel(file_path)

# Create a new Excel workbook
wb = Workbook()

# Add the first sheet (the default sheet)
wb.remove(wb.active)

# Iterate through each row in the dataframe and create a new sheet for each
for index, row in df.iterrows():
    sheet_name = f"Row_{index + 1}"  # Name the sheet based on the row number
    sheet = wb.create_sheet(sheet_name)

    # Write the column titles (first row)
    for col_num, col_name in enumerate(df.columns, 1):
        sheet.cell(row=col_num, column=1, value=col_name)

    # Write the data row
    for col_num, value in enumerate(row, 1):
        sheet.cell(row=col_num, column=2, value=value)

# Save the new Excel file with 20 sheets
wb.save("output_excel_file.xlsx")
